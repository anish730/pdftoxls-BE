from flask import Flask, request, jsonify, send_from_directory
import os
import mimetypes
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
import csv
import re
import logging
import traceback
from werkzeug.utils import secure_filename
from logging.handlers import RotatingFileHandler
from flasgger import Swagger, swag_from
from flask_cors import CORS
from table_extractor import TableExtractor

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
CONVERTED_FILES_FOLDER = 'converted_files'  # New folder for converted files
ALLOWED_EXTENSIONS = {'pdf'}  # Restrict to PDF only
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB max file size

# Define allowed download extensions and their MIME types
ALLOWED_DOWNLOAD_EXTENSIONS = {
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'csv': 'text/csv',
    'pdf': 'application/pdf'
}

# Swagger configuration
swagger_config = {
    "headers": [
        ('Access-Control-Allow-Origin', '*'),
        ('Access-Control-Allow-Methods', 'GET, POST, OPTIONS'),
        ('Access-Control-Allow-Headers', 'Content-Type'),
    ],
    "specs": [
        {
            "endpoint": 'apispec',
            "route": '/apispec.json',
            "rule_filter": lambda rule: True,
            "model_filter": lambda tag: True,
            "version": "2.0"  # Explicitly set Swagger version
        }
    ],
    "static_url_path": "/flasgger_static",
    "swagger_ui": True,
    "specs_route": "/docs",
    "url_prefix": ""
}

# Swagger template
template = {
    "swagger": "2.0",  # Swagger version
    "info": {
        "title": "PDF Conversion API",
        "description": "API for converting PDF files to Excel and CSV formats",
        "contact": {
            "name": "API Support",
            "url": "http://www.yourcompany.com/support",
            "email": "support@yourcompany.com"
        },
        "version": "1.0.0"
    },
    "host": "localhost:5000",  # Host and port
    "basePath": "/",  # Base path prefix
    "schemes": ["http"],  # Protocol schemes
    "consumes": ["application/json", "multipart/form-data"],
    "produces": ["application/json"],
    "definitions": {  # Schema definitions (renamed from components)
        "Error": {
            "type": "object",
            "properties": {
                "error": {"type": "string"},
                "details": {"type": "string"},
                "error_id": {"type": "string"}
            }
        },
        "SuccessResponse": {
            "type": "object",
            "properties": {
                "message": {"type": "string"},
                "filename": {"type": "string"},
                "file_size": {"type": "integer"}
            }
        }
    },
    "securityDefinitions": {  # Security definitions (renamed from securitySchemes)
        "ApiKeyAuth": {
            "type": "apiKey",
            "in": "header",
            "name": "X-API-Key"
        }
    },
    "security": [  # Default security requirement
        {"ApiKeyAuth": []}
    ]
}


# Configure logging
def setup_logging():
    # Create logs directory if it doesn't exist
    os.makedirs('logs', exist_ok=True)
    
    # Configure logging format
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # File handler with rotation
    file_handler = RotatingFileHandler(
        'logs/app.log',
        maxBytes=1024 * 1024,  # 1MB
        backupCount=10
    )
    file_handler.setFormatter(formatter)
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    
    # Get the root logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Remove existing handlers to avoid duplicates
    logger.handlers = []
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


# Initialize logging
logger = setup_logging()

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['CONVERTED_FILES_FOLDER'] = CONVERTED_FILES_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Initialize Swagger with template and config
swagger = Swagger(app, template=template, config=swagger_config)


# Add CORS headers to all responses
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response


class ConversionError(Exception):
    """Custom exception for conversion-related errors"""
    pass


class ValidationError(Exception):
    """Custom exception for validation-related errors"""
    pass


def handle_error(e: Exception, context: str = "") -> tuple:
    """
    Centralized error handling function that logs errors and returns appropriate responses
    """
    error_id = datetime.now().strftime('%Y%m%d%H%M%S')
    
    if isinstance(e, ValidationError):
        logger.warning(f"Validation error in {context}: {str(e)}")
        return {
            'error': 'Validation Error',
            'details': str(e),
            'error_id': error_id
        }, 400
        
    elif isinstance(e, ConversionError):
        logger.error(f"Conversion error in {context}: {str(e)}")
        return {
            'error': 'Conversion Error',
            'details': str(e),
            'error_id': error_id
        }, 500
        
    elif isinstance(e, FileNotFoundError):
        logger.error(f"File not found in {context}: {str(e)}")
        return {
            'error': 'File Not Found',
            'details': str(e),
            'error_id': error_id
        }, 404
        
    else:
        # Log the full traceback for unexpected errors
        logger.error(f"Unexpected error in {context}: {str(e)}\n{traceback.format_exc()}")
        return {
            'error': 'Internal Server Error',
            'details': 'An unexpected error occurred. Please contact support with the error ID.',
            'error_id': error_id
        }, 500


#test
def allowed_file(filename):
    try:
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    except Exception as e:
        logger.error(f"Error checking allowed file: {str(e)}")
        return False


def is_pdf(file):
    try:
        if not file or not file.filename:
            raise ValidationError("No file provided")
        
        # Check file extension
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext != '.pdf':
            raise ValidationError("File must have .pdf extension")
        
        # Check MIME type
        try:
            file_content = file.read(1024)
            file.seek(0)
            mime_type = mimetypes.guess_type(file.filename)[0]
            if mime_type != 'application/pdf':
                raise ValidationError("Invalid PDF MIME type")
        except Exception as e:
            raise ValidationError(f"Error checking MIME type: {str(e)}")
            
        return True
        
    except ValidationError as e:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in is_pdf: {str(e)}")
        return False


def extract_text_from_page(page, extraction_method="blocks"):
    """
    Extract text from a page using different methods based on content structure.
    Returns a list of text blocks/lines with their positions.
    """
    if extraction_method == "blocks":
        # Get text in blocks - good for maintaining paragraph structure
        blocks = page.get_text("blocks")
        # Sort blocks by vertical position (y0), then horizontal position (x0)
        sorted_blocks = sorted(blocks, key=lambda b: (b[1], b[0]))
        return [(block[4], (block[0], block[1])) for block in sorted_blocks]
    
    elif extraction_method == "words":
        # Get individual words with positions - good for maintaining word order
        words = page.get_text("words")
        sorted_words = sorted(words, key=lambda w: (w[3], w[0]))  # Sort by y, then x
        return [(word[4], (word[0], word[3])) for word in sorted_words]
    
    elif extraction_method == "html":
        # Get HTML structure - good for tables and formatted text
        html = page.get_text("html")
        # Basic HTML parsing (you might want to use BeautifulSoup for better parsing)
        return [(html, (0, 0))]
    
    else:  # "text"
        # Simple text extraction - fallback method
        return [(page.get_text("text"), (0, 0))]


def detect_content_type(page):
    """
    Analyze page content to determine the best extraction method.
    """
    # Get text in different formats for analysis
    blocks = page.get_text("blocks")
    words = page.get_text("words")
    
    # Check for potential table structure
    if len(blocks) > 0:
        # Calculate average block width and height
        avg_width = sum(b[2] - b[0] for b in blocks) / len(blocks)
        avg_height = sum(b[3] - b[1] for b in blocks) / len(blocks)
        
        # If blocks are regularly sized and aligned, might be a table
        regular_sizes = all(
            abs((b[2] - b[0]) - avg_width) < 20 and 
            abs((b[3] - b[1]) - avg_height) < 20 
            for b in blocks
        )
        
        if regular_sizes and len(blocks) > 5:
            return "blocks"  # Likely a table structure
    
    # Check word density and distribution
    if len(words) > 100:  # Dense text
        return "words"
    
    # Default to simple text for basic content
    return "text"


def format_excel_sheet(sheet, headers=True):
    """
    Apply formatting to the Excel sheet for better readability.
    """
    # Format headers
    if headers:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
        sheet.column_dimensions[column_letter].width = adjusted_width

def extract_table_data(pdf_path):
    """
    Extract tabular data from a PDF file.
    Returns a list of tables, where each table is a list of rows.
    Each row is a list of cell values.
    """
    try:
        # Open the PDF
        doc = fitz.open(pdf_path)
        all_tables = []
        
        # Process each page
        for page_num in range(len(doc)):
            page = doc[page_num]
            
            # Find tables on the page
            tables = page.find_tables()
            if not tables:
                continue
                
            # Process each table found on the page
            for table in tables:
                table_data = []
                
                # Extract cells from the table
                for row_cells in table.extract():
                    # Clean and process each cell in the row
                    row_data = [
                        cell.strip() if isinstance(cell, str) else str(cell)
                        for cell in row_cells
                        if cell is not None
                    ]
                    
                    # Only add non-empty rows
                    if any(cell for cell in row_data):
                        table_data.append(row_data)
                
                # Only add tables that have data
                if table_data:
                    all_tables.append(table_data)
        
        return all_tables
        
    except Exception as e:
        logger.error(f"Error extracting tables from PDF: {str(e)}")
        raise Exception(f"Failed to extract tables: {str(e)}")

def convert_tables_to_excel(tables, output_path):
    """
    Convert extracted tables to Excel format with proper formatting.
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tables"
        
        current_row = 1
        table_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Process each table
        for table_num, table in enumerate(tables, 1):
            # Add table header/separator
            header = f"Table {table_num}"
            sheet.cell(row=current_row, column=1, value=header)
            sheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            if not table:
                continue
                
            # Get the maximum number of columns in this table
            max_cols = max(len(row) for row in table)
            
            # Write table data
            for row_num, row in enumerate(table):
                for col_num, cell_value in enumerate(row, 1):
                    cell = sheet.cell(row=current_row, column=col_num)
                    cell.value = cell_value
                    cell.border = table_border
                    
                    # Format header row
                    if row_num == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    cell.alignment = Alignment(wrap_text=True)
                
                current_row += 1
            
            # Add spacing between tables
            current_row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(output_path)
        return output_path
        
    except Exception as e:
        logger.error(f"Error converting tables to Excel: {str(e)}")
        raise Exception(f"Failed to convert tables to Excel: {str(e)}")

def convert_tables_to_csv(tables, output_path):
    """
    Convert extracted tables to CSV format.
    """
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
            
            # Process each table
            for table_num, table in enumerate(tables, 1):
                # Add table header/separator
                writer.writerow([f"Table {table_num}"])
                
                if not table:
                    continue
                
                # Write table data
                for row in table:
                    writer.writerow(row)
                
                # Add blank row between tables
                writer.writerow([])
        
        return output_path
        
    except Exception as e:
        logger.error(f"Error converting tables to CSV: {str(e)}")
        raise Exception(f"Failed to convert tables to CSV: {str(e)}")

def convert_pdf_to_excel(pdf_path):
    """
    Convert a PDF file to Excel format, focusing on bank statement data.
    Creates a single continuous sheet of transaction records.
    """
    try:
        # Initialize table extractor
        extractor = TableExtractor()
        
        # Extract tables from PDF
        df = extractor.extract_tables(pdf_path)
        
        if df.empty:
            # Fallback to original text extraction if no tables found
            logger.warning("No tables found, falling back to text extraction")
            return convert_pdf_to_excel_fallback(pdf_path)
        
        # Ensure converted files directory exists
        os.makedirs(CONVERTED_FILES_FOLDER, exist_ok=True)
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = os.path.splitext(os.path.basename(pdf_path))[0]
        excel_path = os.path.join(
            CONVERTED_FILES_FOLDER,
            f"{base_filename}_{timestamp}.xlsx"
        )
        
        # Create Excel workbook with better formatting
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transaction Records"
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        amount_alignment = Alignment(horizontal='right', vertical='center')
        date_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = header_alignment
        
        # Write data with appropriate formatting
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                # Apply column-specific formatting
                column_name = df.columns[col_idx - 1]
                if 'Date' in column_name:
                    cell.alignment = date_alignment
                elif any(x in column_name for x in ['Withdrawals', 'Deposits', 'Balance', 'Amount']):
                    cell.alignment = amount_alignment
                    # Format as currency if it's a number
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = text_alignment
        
        # Auto-adjust column widths with minimum and maximum values
        min_width = 10
        max_width = 50
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Find the maximum length in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Apply width constraints
            adjusted_width = max(min(max_length + 2, max_width), min_width)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        sheet.freeze_panes = 'A2'
        
        # Save the workbook
        workbook.save(excel_path)
        return excel_path
        
    except Exception as e:
        logger.error(f"Error in PDF to Excel conversion: {str(e)}")
        raise Exception(f"Failed to convert PDF to Excel: {str(e)}")

def convert_pdf_to_csv(pdf_path):
    """
    Convert a PDF file to CSV format, focusing on bank statement data.
    Creates a single continuous CSV file of transaction records.
    """
    try:
        # Initialize table extractor
        extractor = TableExtractor()
        
        # Extract tables from PDF
        df = extractor.extract_tables(pdf_path)
        
        if df.empty:
            # Fallback to original text extraction if no tables found
            logger.warning("No tables found, falling back to text extraction")
            return convert_pdf_to_csv_fallback(pdf_path)
        
        # Ensure converted files directory exists
        os.makedirs(CONVERTED_FILES_FOLDER, exist_ok=True)
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = os.path.splitext(os.path.basename(pdf_path))[0]
        csv_path = os.path.join(
            CONVERTED_FILES_FOLDER,
            f"{base_filename}_{timestamp}.csv"
        )
        
        # Save to CSV with proper formatting
        df.to_csv(
            csv_path,
            index=False,
            encoding='utf-8-sig',  # Include BOM for Excel compatibility
            quoting=csv.QUOTE_ALL,  # Quote all fields
            date_format='%Y-%m-%d',  # Consistent date format
            float_format='%.2f'  # Format numbers with 2 decimal places
        )
        
        return csv_path
        
    except Exception as e:
        logger.error(f"Error in PDF to CSV conversion: {str(e)}")
        raise Exception(f"Failed to convert PDF to CSV: {str(e)}")

# Rename original conversion functions as fallback methods
convert_pdf_to_excel_fallback = convert_pdf_to_excel
convert_pdf_to_csv_fallback = convert_pdf_to_csv

def clean_text_for_csv(text):
    """
    Clean and prepare text for CSV formatting.
    Handles special characters, multiple spaces, and newlines.
    """
    # Replace multiple spaces with single space
    text = re.sub(r'\s+', ' ', text)
    # Remove special characters that might cause issues
    text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
    # Trim whitespace
    return text.strip()


def extract_structured_text(page):
    """
    Extract text from a page in a structured way, attempting to preserve
    natural text boundaries and table-like structures.
    Returns a list of text rows.
    """
    rows = []
    current_row = []
    last_y = None
    y_threshold = 5  # pixels

    # Get blocks of text with their positions
    blocks = page.get_text("blocks")
    # Sort blocks by vertical position first, then horizontal
    sorted_blocks = sorted(blocks, key=lambda b: (b[1], b[0]))

    for block in sorted_blocks:
        text = clean_text_for_csv(block[4])
        if not text:
            continue

        y_pos = block[1]  # y-coordinate of the block

        # If this is a new row (based on y-position)
        if last_y is None or abs(y_pos - last_y) > y_threshold:
            if current_row:
                rows.append(current_row)
            current_row = [text]
            last_y = y_pos
        else:
            # Same row, add as new column
            current_row.append(text)

    # Don't forget the last row
    if current_row:
        rows.append(current_row)

    return rows

@app.route('/')
def hello_world():
    return jsonify({
        'message': 'PDF Conversion API',
        'version': '1.0.0',
        'status': 'running'
    })


@app.route('/upload', methods=['POST'])
@swag_from({
    'tags': ['Upload'],
    'summary': 'Upload a PDF file',
    'description': 'Upload a PDF file to the server for later conversion',
    'consumes': ['multipart/form-data'],
    'produces': ['application/json'],
    'parameters': [
        {
            'name': 'file',
            'in': 'formData',
            'type': 'file',
            'required': True,
            'description': 'PDF file to upload'
        }
    ],
    'responses': {
        '200': {
            'description': 'File successfully uploaded',
            'schema': {
                'type': 'object',
                'properties': {
                    'message': {'type': 'string'},
                    'filename': {'type': 'string'},
                    'file_size': {'type': 'integer'}
                }
            }
        },
        '400': {
            'description': 'Invalid input',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        },
        '500': {
            'description': 'Server error',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        }
    }
})
def upload_file():
    try:
        logger.info("Starting file upload process")
        
        # Check if the post request has the file part
        if 'file' not in request.files:
            raise ValidationError("No file part in the request")

        file = request.files['file']

        # Check if a file was selected
        if file.filename == '':
            raise ValidationError("No selected file")

        # Check if it's a PDF file
        if not is_pdf(file):
            raise ValidationError("Invalid or corrupted PDF file")

        # If file is valid, save it
        filename = secure_filename(file.filename)
        logger.info(f"Processing upload for file: {filename}")
        
        # Ensure upload directory exists
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # Save the file
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(pdf_path)

        # Verify file was saved successfully
        if not os.path.exists(pdf_path):
            raise ValidationError("Failed to save uploaded file")
            
        response = {
            'message': 'File uploaded successfully',
            'filename': filename,
            'file_size': os.path.getsize(pdf_path)
        }
        
        logger.info(f"Successfully uploaded file: {filename}")
        return response, 200

    except Exception as e:
        return handle_error(e, "upload_file")


@app.route('/convert', methods=['POST'])
@swag_from({
    'tags': ['Conversion'],
    'summary': 'Convert an uploaded PDF file',
    'description': 'Convert a previously uploaded PDF file to either Excel or CSV format',
    'consumes': ['application/json'],
    'produces': ['application/json'],
    'parameters': [
        {
            'name': 'body',
            'in': 'body',
            'required': True,
            'schema': {
                'type': 'object',
                'properties': {
                    'filename': {
                        'type': 'string',
                        'description': 'Name of the PDF file to convert'
                    },
                    'formats': {
                        'type': 'array',
                        'items': {
                            'type': 'string',
                            'enum': ['excel', 'csv']
                        },
                        'description': 'List of desired output formats',
                        'minItems': 1
                    }
                },
                'required': ['filename', 'formats']
            }
        }
    ],
    'responses': {
        '200': {
            'description': 'File successfully converted',
            'schema': {
                'type': 'object',
                'properties': {
                    'message': {'type': 'string'},
                    'input_file': {'type': 'string'},
                    'conversions': {
                        'type': 'array',
                        'items': {
                            'type': 'object',
                            'properties': {
                                'format': {'type': 'string'},
                                'output_file': {'type': 'string'},
                                'file_size': {'type': 'integer'},
                                'download_url': {'type': 'string'}
                            }
                        }
                    }
                }
            }
        },
        '400': {
            'description': 'Invalid input',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        },
        '404': {
            'description': 'File not found',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        },
        '500': {
            'description': 'Server error',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        }
    }
})
def convert_file():
    try:
        logger.info("Starting conversion request")
        
        # Get parameters from request
        data = request.get_json() if request.is_json else request.form
        
        # Check if filename is provided
        filename = data.get('filename')
        if not filename:
            raise ValidationError("No filename provided")
            
        # Check if formats are provided
        formats = data.get('formats', [])
        if not formats:
            raise ValidationError("No output formats specified")
            
        # Validate formats
        invalid_formats = [fmt for fmt in formats if fmt not in ['excel', 'csv']]
        if invalid_formats:
            raise ValidationError(f"Invalid formats: {', '.join(invalid_formats)}. Must be 'excel' or 'csv'")
            
        logger.info(f"Converting {filename} to formats: {formats}")
            
        # Construct the full path to the PDF file
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Check if the file exists
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"File {filename} not found")
            
        # Store conversion results
        conversions = []
            
        try:
            # Convert the file based on requested formats
            if 'excel' in formats:
                excel_path = convert_pdf_to_excel(pdf_path)
                conversions.append({
                    'format': 'excel',
                    'output_file': os.path.basename(excel_path),
                    'file_size': os.path.getsize(excel_path),
                    'download_url': f"/download/{os.path.basename(excel_path)}"
                })
                
            if 'csv' in formats:
                csv_path = convert_pdf_to_csv(pdf_path)
                conversions.append({
                    'format': 'csv',
                    'output_file': os.path.basename(csv_path),
                    'file_size': os.path.getsize(csv_path),
                    'download_url': f"/download/{os.path.basename(csv_path)}"
                })
                
            response = {
                'message': 'File converted successfully',
                'input_file': filename,
                'conversions': conversions
            }
            
            logger.info(f"Successfully converted {filename}")
            return response, 200
            
        except Exception as e:
            raise ConversionError(f"Conversion failed: {str(e)}")
            
    except Exception as e:
        return handle_error(e, "convert_file")


def is_safe_file_to_download(filename):
    """
    Check if the file is safe to download based on its extension and existence.
    """
    if not filename:
        return False
    
    # Check file extension
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    if ext not in ALLOWED_DOWNLOAD_EXTENSIONS:
        return False
    
    # Ensure the filename is secure
    secure_name = secure_filename(filename)
    if secure_name != filename:
        return False
    
    return True


@app.route('/download/<filename>')
@swag_from({
    'tags': ['Download'],
    'summary': 'Download a converted file',
    'description': 'Download a previously converted Excel or CSV file',
    'produces': [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'text/csv'
    ],
    'parameters': [
        {
            'name': 'filename',
            'in': 'path',
            'required': True,
            'type': 'string',
            'description': 'Name of the file to download'
        }
    ],
    'responses': {
        '200': {
            'description': 'File download',
            'schema': {
                'type': 'file'
            }
        },
        '400': {
            'description': 'Invalid input',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        },
        '404': {
            'description': 'File not found',
            'schema': {
                '$ref': '#/definitions/Error'
            }
        }
    }
})
def download_file(filename):
    try:
        logger.info(f"Processing download request for {filename}")
        
        # Security checks
        if not filename:
            raise ValidationError("No filename provided")
            
        if not is_safe_file_to_download(filename):
            raise ValidationError("File type not allowed for download")
        
        # Get file extension and check if it's allowed
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        if file_ext not in ALLOWED_DOWNLOAD_EXTENSIONS:
            raise ValidationError(f"Files with extension .{file_ext} are not allowed")
            
        # Check if file exists in converted files directory
        file_path = os.path.join(CONVERTED_FILES_FOLDER, filename)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {filename} not found")
            
        # Get MIME type
        mime_type = ALLOWED_DOWNLOAD_EXTENSIONS.get(file_ext)
        
        # Set additional headers for different file types
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        # If it's a CSV file, add UTF-8 BOM for Excel compatibility
        if file_ext == 'csv':
            headers['Content-Type'] = f'{mime_type}; charset=utf-8-sig'
        
        logger.info(f"Sending file: {filename}")
        
        return send_from_directory(
            CONVERTED_FILES_FOLDER,
            filename,
            mimetype=mime_type,
            as_attachment=True,
            download_name=filename,
            etag=True,
            max_age=0,
            conditional=True,
            headers=headers
        )
        
    except Exception as e:
        return handle_error(e, "download_file")


# Error handlers for common HTTP errors
@app.errorhandler(404)
def not_found_error(error):
    logger.warning(f"404 error: {request.url}")
    return jsonify({
        'error': 'Not Found',
        'details': 'The requested resource was not found'
    }), 404


@app.errorhandler(405)
def method_not_allowed_error(error):
    logger.warning(f"405 error: {request.method} {request.url}")
    return jsonify({
        'error': 'Method Not Allowed',
        'details': 'The method is not allowed for the requested URL'
    }), 405


@app.errorhandler(413)
def request_entity_too_large_error(error):
    logger.warning(f"413 error: File too large")
    return jsonify({
        'error': 'File Too Large',
        'details': f'The file exceeds the maximum allowed size of {MAX_FILE_SIZE/1024/1024}MB'
    }), 413


if __name__ == '__main__':
    logger.info("Starting Flask application")
    app.run(debug=True)
